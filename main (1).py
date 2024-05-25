import pandas as pd
import requests
from bs4 import BeautifulSoup
import os

input_file = pd.read_excel("Input.xlsx")
input_file_shape = input_file.shape
input_index_max = input_file_shape[0]
input_index = 0

if not os.path.exists('Extracted_Text_Files'):
            os.makedirs('Extracted_Text_Files')

while input_index < input_index_max:
    url = input_file.URL[input_index]
    url_id = input_file.URL_ID[input_index]
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        if soup.find('h1'):
            heading = soup.find('h1').get_text()
        else:
            heading = ""
        paragraphs = soup.find_all('p')
        text = ' '.join([p.get_text() for p in paragraphs])
        with open(f"Extracted_Text_Files/{url_id}.txt", 'w', encoding='utf-8') as file:
            file.write(f"{heading}\n\n")
            file.write(f"{text}\n")
        print(f"Saved content from given URL {url_id} to Extracted_Text_Files/{url_id}.txt")
    else:
        with open(f"Extracted_Text_Files/{url_id}.txt", 'w', encoding='utf-8') as file:
            file.write(f"Error in retrieving data from {url_id}\n\n.")
        print(f"Error in retrieving content from given URL {url_id}")
    input_index += 1

StopWordFiles = ['StopWords_Auditor.txt', 'StopWords_Currencies.txt', 'StopWords_DatesandNumbers.txt',
                 'StopWords_Generic.txt', 'StopWords_GenericLong.txt',
                 'StopWords_Geographic.txt', 'StopWords_Names.txt']
with open('StopWords_Master.txt', 'w') as output:
    for file_name in StopWordFiles:
        with open(file_name, 'r') as StopWordFile:
            content = StopWordFile.read()
            output.write(content)
            output.write('\n')


def TextCleaning(ExtractedFile, StopWordsList, CleanedFile):
    with open(ExtractedFile, 'r', encoding='utf-8') as file:
        extracted_content = file.read().lower()
    with open(StopWordsList, 'r') as file:
        StopTerms = file.read().lower().split()
    tokens = wordpunct_tokenize(extracted_content)
    Cleaned_Words = [word for word in tokens if word not in StopTerms]
    Cleaned_Content = ' '.join(Cleaned_Words)
    with open(CleanedFile, 'w', encoding='utf-8') as file:
        file.write(Cleaned_Content)


def WordCount(ExtractedFile, StopWordsList):
    with open(ExtractedFile, 'r', encoding='utf-8') as file:
        extracted_content = file.read().lower()
    with open(StopWordsList, 'r') as file:
        StopTerms = file.read().lower().split()
    tokens = wordpunct_tokenize(extracted_content)
    Cleaned_Words = [word for word in tokens if word not in StopTerms]
    word_count = len(Cleaned_Words)
    return word_count

def WordCountTotal(ExtractedFile):
    with open(ExtractedFile, 'r', encoding='utf-8') as file:
        extracted_content = file.read().lower()
    tokens = wordpunct_tokenize(extracted_content)
    word_count = len(tokens)
    return word_count


def DictionaryCleaning(ExtractedFile, StopWordsList, CleanedFile):
    with open(ExtractedFile, 'r') as file:
        extracted_content = file.read().lower()
    with open(StopWordsList, 'r') as file:
        StopTerms = file.read().lower().split()
    tokens = wordpunct_tokenize(extracted_content)
    Cleaned_Words = [word for word in tokens if word not in StopTerms]
    Cleaned_Content = ' '.join(Cleaned_Words)
    with open(CleanedFile, 'w', encoding='utf-8') as file:
        file.write(Cleaned_Content)

import nltk
from nltk.tokenize import wordpunct_tokenize


def PositiveScore(CleanedFile, PositiveDictionary):
    positive_score = 0
    with open(CleanedFile, 'r', encoding='utf-8') as file:
        cleaned_content = file.read().lower()
    with open(PositiveDictionary, 'r') as file:
        PositiveTerms = file.read().lower().split()
    tokens = wordpunct_tokenize(cleaned_content)
    for token in tokens:
        if token in PositiveTerms:
            positive_score += 1
    return positive_score

def NegativeScore(CleanedFile, NegativeDictionary):
    negative_score = 0
    with open(CleanedFile, 'r', encoding='utf-8') as file:
        cleaned_content = file.read().lower()
    with open(NegativeDictionary, 'r') as file:
        NegativeTerms = file.read().lower().split()
    tokens = wordpunct_tokenize(cleaned_content)
    for token in tokens:
        if token in NegativeTerms:
            negative_score += 1
    return negative_score

def Sentences(ExtractedFile):
    sentences = 0
    punctuation = [".", "!", "?"]
    with open(ExtractedFile, 'r', encoding='utf-8') as file:
        content = file.read().lower()
    for char in content:
        if char in punctuation:
            sentences += 1
    return sentences


import nltk
from nltk.corpus import cmudict
from nltk.tokenize import word_tokenize
nltk.download('cmudict')
dictionary = cmudict.dict()


def count_complex_words(ExtractedFile):
    complex_word_count = 0
    with open(ExtractedFile, 'r', encoding='utf-8') as file:
        text = file.read().lower()
        words = word_tokenize(text)
        for word in words:
            if word in dictionary:
                syllables = [len(list(y for y in x if y[-1].isdigit())) for x in dictionary[word]][0]
                if syllables > 2:
                    complex_word_count += 1
    return complex_word_count

import re


def count_syllables_in_text(ExtractedFile):
    word_pattern = r'\b\w+\b'
    total_syllables = 0
    with open(ExtractedFile, 'r', encoding='utf-8') as file:
        text = file.read().lower()
        words = re.findall(word_pattern, text)
        for word in words:
            pronunciation = dictionary.get(word)
            if pronunciation:
                syllable_count = len(list(filter(lambda s: s[-1].isdigit(), pronunciation[0])))
                total_syllables += syllable_count
    return total_syllables

def personalpronouns(ExtractedFile):
    personal_pronouns = 0
    with open(ExtractedFile, 'r', encoding='utf-8') as file:
        text = file.read()
        words = word_tokenize(text)
        pronouns = ["I", "we", "my", "ours", "us", "We", "My", "Ours", "Us"]
        for word in words:
            if word in pronouns:
                personal_pronouns += 1
    return personal_pronouns


import string

def count_characters(ExtractedFile):
    punctuation = string.punctuation
    count = 0
    with open(ExtractedFile, 'r', encoding='utf-8') as file:
        text = file.read()
        for char in text:
            if char not in string.whitespace and char not in punctuation:
                count += 1
    return count


def PolarityScore(positive_score, negative_score):
    polarity_score = (positive_score - negative_score)/((positive_score + negative_score) + 0.000001)
    return polarity_score


def SubjectivityScore(positive_score, negative_score, word_count):
    subjectivity_score = (positive_score + negative_score)/(word_count + 0.000001)
    return subjectivity_score


if not os.path.exists('Cleaned_Text_Files'):
            os.makedirs('Cleaned_Text_Files')

ExtractedFile = 'positive-words.txt'
StopWordsList = 'StopWords_Master.txt'
PositiveDictionary = 'positive_final.txt'
DictionaryCleaning(ExtractedFile, StopWordsList, PositiveDictionary)
print("Positive Words Dictionary has been prepared.")
ExtractedFile = 'negative-words.txt'
StopWordsList = 'StopWords_Master.txt'
NegativeDictionary = 'negative_final.txt'
DictionaryCleaning(ExtractedFile, StopWordsList, NegativeDictionary)
print("Negative Words Dictionary has been prepared.")

from openpyxl import Workbook
headings = ["URL_ID", "URL", "POSITIVE SCORE", "NEGATIVE SCORE", "POLARITY SCORE", "SUBJECTIVITY SCORE",
            "AVG SENTENCE LENGTH", "PERCENTAGE OF COMPLEX WORDS", "FOG INDEX", "AVG NUMBER OF WORDS PER SENTENCE",
            "COMPLEX WORD COUNT", "WORD COUNT", "SYLLABLE PER WORD", "PERSONAL PRONOUNS", "AVG WORD LENGTH"]
workbook = Workbook()
worksheet = workbook.active
for col, heading in enumerate(headings, start=1):
    worksheet.cell(row=1, column=col, value=heading)
workbook.save("Output.xlsx")
worksheet.title = "Sheet"


ExcelFile = "Output.xlsx"
worksheet_name = "Sheet"


import openpyxl


def excel(excelfile, output, data):
    workbook = openpyxl.load_workbook(excelfile)
    sheet = workbook[output]
    next_row = sheet.max_row + 1
    for col_num, cell_value in enumerate(data, start=1):
        sheet.cell(row=next_row, column=col_num, value=cell_value)
    workbook.save(excelfile)
    print(f"Data row written to {excelfile} successfully!")

input_index = 0

while input_index < input_index_max:
    url = input_file.URL[input_index]
    url_id = input_file.URL_ID[input_index]
    ExtractedFile = f"Extracted_Text_Files/{url_id}.txt"
    StopWordsList = 'StopWords_Master.txt'
    CleanedFile = f"Cleaned_Text_Files/Cleaned_{url_id}.txt"
    with open(ExtractedFile, 'r', encoding='utf-8', errors='ignore') as file:
        content = file.read().lower()
        if content == f"Error in retrieving data from {url_id}\n\n":
            with open(CleanedFile, 'w', encoding='utf-8') as file:
                file.write("URL Error: No Data Retrieved")
        else:
            TextCleaning(ExtractedFile, StopWordsList, CleanedFile)
            print(f"Text file for {url_id} is cleaned.")
            positive_score = PositiveScore(CleanedFile, PositiveDictionary)
            negative_score = NegativeScore(CleanedFile, NegativeDictionary)
            polarity_score = PolarityScore(positive_score, negative_score)
            word_count = WordCount(ExtractedFile, StopWordsList)
            subjectivity_score = SubjectivityScore(positive_score, negative_score, word_count)
            sentences = Sentences(ExtractedFile)
            complex_words = count_complex_words(ExtractedFile)
            word_count_total = WordCountTotal(ExtractedFile)
            avg_word_sen = (word_count_total/sentences)
            per_complex = (complex_words/word_count_total)
            avg_sen_len = (word_count_total/sentences)
            fog_index = 0.4 * (avg_sen_len + per_complex)
            personal_pronouns = personalpronouns(ExtractedFile)
            syllables = count_syllables_in_text(ExtractedFile)
            syllable_per_word = syllables/word_count_total
            char_count = count_characters(ExtractedFile)
            avg_wordlength = char_count/word_count_total
            data = [url_id, url, positive_score, negative_score, polarity_score, subjectivity_score, avg_sen_len,
                    per_complex, fog_index, avg_word_sen, complex_words, word_count, syllable_per_word,
                    personal_pronouns, avg_wordlength]
            if data[11] == 7:
                data_err = [url_id, url, "Error retrieving data from URL"]
                excel(ExcelFile, worksheet_name, data_err)
            else:
                excel(ExcelFile, worksheet_name, data)
            input_index += 1

print("\n\ncomplete.")







